"""Utilidades para generar reportes empresariales."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Dict, List, Optional

from django.db.models import Count, Sum
from django.db.models.functions import TruncMonth, TruncYear
from django.utils import timezone

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from .models import (
    Bitacora,
    Computadora,
    Impresora,
    Insumo,
    Monitor,
    Networking,
    Periferico,
    Software,
    TecnologiaMedica,
    Telefonia,
)


@dataclass(frozen=True)
class AssetConfig:
    key: str
    label: str
    model: type
    acquisition_field: Optional[str] = None
    warranty_field: Optional[str] = None
    value_field: Optional[str] = None
    license_field: Optional[str] = None
    segment: str = "informatica"


SEGMENT_INFORMATICA = "informatica"
SEGMENT_MEDICA = "medica"


BITACORA_SEGMENT_TYPES = {
    SEGMENT_INFORMATICA: {
        "computadora",
        "impresora",
        "monitor",
        "networking",
        "telefonia",
        "periferico",
        "software",
        "insumo",
    },
    SEGMENT_MEDICA: {"tecnologia_medica"},
}


ASSET_CONFIGS: List[AssetConfig] = [
    AssetConfig(
        key="computadora",
        label="Computadoras",
        model=Computadora,
        acquisition_field="fecha_adquisicion",
        warranty_field="fecha_finalizacion_garantia",
        value_field="valor_adquisicion",
        segment=SEGMENT_INFORMATICA,
    ),
    AssetConfig(
        key="impresora",
        label="Impresoras",
        model=Impresora,
        acquisition_field="fecha_adquisicion",
        warranty_field="fecha_finalizacion_garantia",
        value_field="valor_adquisicion",
        segment=SEGMENT_INFORMATICA,
    ),
    AssetConfig(
        key="monitor",
        label="Monitores",
        model=Monitor,
        acquisition_field="fecha_adquisicion",
        warranty_field="fecha_finalizacion_garantia",
        value_field="valor_adquisicion",
        segment=SEGMENT_INFORMATICA,
    ),
    AssetConfig(
        key="networking",
        label="Networking",
        model=Networking,
        acquisition_field="fecha_adquisicion",
        warranty_field="fecha_finalizacion_garantia",
        value_field="valor_adquisicion",
        segment=SEGMENT_INFORMATICA,
    ),
    AssetConfig(
        key="telefonia",
        label="Telefonía",
        model=Telefonia,
        acquisition_field="fecha_adquisicion",
        warranty_field="fecha_finalizacion_garantia",
        value_field="valor_adquisicion",
        segment=SEGMENT_INFORMATICA,
    ),
    AssetConfig(
        key="periferico",
        label="Periféricos",
        model=Periferico,
        acquisition_field="fecha_adquisicion",
        warranty_field="fecha_finalizacion_garantia",
        value_field="valor_adquisicion",
        segment=SEGMENT_INFORMATICA,
    ),
    AssetConfig(
        key="insumo",
        label="Insumos",
        model=Insumo,
        acquisition_field=None,
        value_field=(
            "valor_unitario_estandar"
            if hasattr(Insumo, "valor_unitario_estandar")
            else None
        ),
        segment=SEGMENT_INFORMATICA,
    ),
    AssetConfig(
        key="software",
        label="Software",
        model=Software,
        acquisition_field="fecha_adquisicion",
        license_field="fecha_expiracion",
        value_field="costo_total",
        segment=SEGMENT_INFORMATICA,
    ),
    AssetConfig(
        key="tecnologia_medica",
        label="Tecnología Médica",
        model=TecnologiaMedica,
        acquisition_field="fecha_adquisicion",
        warranty_field="fecha_finalizacion_garantia",
        value_field="valor_adquisicion",
        segment=SEGMENT_MEDICA,
    ),
]

OBSOLESCENCE_YEARS = 5
WARRANTY_SOON_THRESHOLD_DAYS = 90


def default_date_range(
    start: Optional[date], end: Optional[date]
) -> Dict[str, date]:
    today = timezone.now().date()
    default_start = today - timedelta(days=365)
    if start and end and start > end:
        start, end = end, start
    return {
        "start": start or default_start,
        "end": end or today,
    }


def _format_currency(value: Optional[Decimal]) -> Decimal:
    if value is None:
        return Decimal(0)
    return value


def build_excel_report(report: Dict[str, object]) -> bytes:
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment, Font, NamedStyle  # type: ignore
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise RuntimeError(
            "openpyxl no está instalado en el entorno actual"
        ) from exc

    workbook = Workbook()

    header_style = NamedStyle(name="header")
    header_style.font = Font(bold=True)
    header_style.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    summary_sheet = workbook.active
    summary_sheet.title = "Resumen"
    summary_sheet.append(["Reporte Empresarial ASSE-GestACT"])
    summary_sheet.append(
        [
            "Período",
            "{} - {}".format(
                report["start_date"].strftime("%d/%m/%Y"),
                report["end_date"].strftime("%d/%m/%Y"),
            ),
        ]
    )
    summary_sheet.append(
        [
            "Generado",
            report["generated_at"]
            .astimezone(timezone.get_current_timezone())
            .strftime("%d/%m/%Y %H:%M"),
        ]
    )
    summary_sheet.append([])
    summary_sheet.append(["Garantías", "Vigentes", "Por vencer", "Vencidas"])
    summary_sheet.append(
        [
            "Totales",
            report["warranty"]["vigentes"],
            report["warranty"]["por_vencer"],
            report["warranty"]["vencidas"],
        ]
    )
    summary_sheet.append([])
    summary_sheet.append(
        [
            "Licencias",
            report["licenses"]["vigentes"],
            report["licenses"]["por_vencer"],
            report["licenses"]["vencidas"],
        ]
    )
    summary_sheet.append([])
    summary_sheet.append(
        [
            "Obsolescencia",
            f">= {report['obsolescence']['threshold_years']} años",
            report["obsolescence"]["obsoletos"],
            report["obsolescence"]["total_evaluados"],
        ]
    )

    for row in summary_sheet.iter_rows(
        min_row=5,
        max_row=7,
        min_col=1,
        max_col=4,
    ):
        for cell in row:
            cell.style = header_style if cell.row == 5 else cell.style

    # Garantías detalle
    warranty_sheet = workbook.create_sheet(title="Garantías")
    warranty_sheet.append(
        ["Activo", "Tipo", "Fecha", "Estado", "Días restantes"]
    )
    for entry in report["warranty"]["detalle"]:
        warranty_sheet.append(
            [
                entry["nombre"],
                entry["tipo"],
                entry["fecha"].strftime("%d/%m/%Y"),
                entry["estado"],
                entry["dias_restantes"],
            ]
        )

    # Licencias detalle
    license_sheet = workbook.create_sheet(title="Licencias")
    license_sheet.append(
        ["Software", "Tipo", "Fecha", "Estado", "Días restantes"]
    )
    for entry in report["licenses"]["detalle"]:
        license_sheet.append(
            [
                entry["nombre"],
                entry["tipo"],
                entry["fecha"].strftime("%d/%m/%Y"),
                entry["estado"],
                entry["dias_restantes"],
            ]
        )

    # Compras anuales
    purchases_sheet = workbook.create_sheet(title="Compras")
    purchases_sheet.append(["Año", "Categoría", "Cantidad", "Monto"])
    for entry in report["annual_purchases"]:
        purchases_sheet.append(
            [
                entry["year"],
                entry["tipo"],
                entry["cantidad"],
                float(_format_currency(entry["monto"])),
            ]
        )

    deliveries_sheet = workbook.create_sheet(title="Entregas")
    deliveries_sheet.append(["Período", "Total (mensual)"])
    for entry in report["monthly_deliveries"]:
        deliveries_sheet.append([entry["periodo"], entry["total"]])
    deliveries_sheet.append([])
    deliveries_sheet.append(["Año", "Total (anual)"])
    for entry in report["annual_deliveries"]:
        deliveries_sheet.append([entry["year"], entry["total"]])

    bajas_sheet = workbook.create_sheet(title="Bajas")
    bajas_sheet.append(["Tipo", "Activo", "Fecha"])
    for entry in report["bajas"]["detalle"]:
        bajas_sheet.append(
            [
                entry["tipo_dispositivo"],
                entry["dispositivo_nombre"],
                entry["fecha_evento"].strftime("%d/%m/%Y %H:%M"),
            ]
        )

    for sheet in workbook.worksheets:
        for column in sheet.columns:
            max_length = max(
                len(str(cell.value)) if cell.value else 0 for cell in column
            )
            column_letter = column[0].column_letter
            sheet.column_dimensions[column_letter].width = max(
                12,
                max_length + 2,
            )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_pdf_report(report: Dict[str, object]) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    story = []

    title = "Reporte Empresarial ASSE-GestACT"
    story.append(Paragraph(title, styles["Title"]))
    subtitle = (
        f"Período: {report['start_date'].strftime('%d/%m/%Y')} - "
        f"{report['end_date'].strftime('%d/%m/%Y')}"
    )
    story.append(Paragraph(subtitle, styles["Normal"]))
    story.append(Spacer(1, 12))

    resumen_data = [
        ["Indicador", "Vigentes", "Por vencer", "Vencidas"],
        [
            "Garantías",
            report["warranty"]["vigentes"],
            report["warranty"]["por_vencer"],
            report["warranty"]["vencidas"],
        ],
        [
            "Licencias",
            report["licenses"]["vigentes"],
            report["licenses"]["por_vencer"],
            report["licenses"]["vencidas"],
        ],
    ]
    resumen_table = Table(resumen_data, hAlign="LEFT")
    resumen_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1d4ed8")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]
        )
    )
    story.append(resumen_table)
    story.append(Spacer(1, 12))

    ob_data = [
        [
            "Tipo",
            "Total",
            "Obsoletos",
            "% Obsolescencia",
            "Antigüedad promedio",
        ]
    ]
    for entry in report["obsolescence"]["por_tipo"]:
        ob_data.append(
            [
                entry["tipo"],
                entry["total"],
                entry["obsoletos"],
                f"{entry['porcentaje']}%",
                f"{entry['antiguedad_promedio']} años",
            ]
        )
    if len(ob_data) > 1:
        ob_table = Table(ob_data, hAlign="LEFT")
        ob_table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor("#0f172a"),
                    ),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ]
            )
        )
        story.append(
            Paragraph("Obsolescencia por categoría", styles["Heading3"])
        )
        story.append(ob_table)
        story.append(Spacer(1, 12))

    if report["annual_purchases"]:
        story.append(Paragraph("Compras por año", styles["Heading3"]))
        pur_data = [["Año", "Categoría", "Cantidad", "Monto"]]
        for entry in report["annual_purchases"]:
            pur_data.append(
                [
                    entry["year"],
                    entry["tipo"],
                    entry["cantidad"],
                    f"${_format_currency(entry['monto']):,.2f}",
                ]
            )
        pur_table = Table(pur_data, hAlign="LEFT")
        pur_table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor("#1e3a8a"),
                    ),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ]
            )
        )
        story.append(pur_table)
        story.append(Spacer(1, 12))

    if report["monthly_deliveries"]:
        story.append(Paragraph("Entregas mensuales", styles["Heading3"]))
        del_data = [["Período", "Total"]]
        for entry in report["monthly_deliveries"]:
            del_data.append([entry["periodo"], entry["total"]])
        del_table = Table(del_data, hAlign="LEFT")
        del_table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor("#0f172a"),
                    ),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ]
            )
        )
        story.append(del_table)
        story.append(Spacer(1, 12))

    if report["bajas"]["total_bajas"]:
        story.append(Paragraph("Bajas registradas", styles["Heading3"]))
        bajas_data = [["Tipo", "Activo", "Fecha"]]
        for entry in report["bajas"]["detalle"]:
            bajas_data.append(
                [
                    entry["tipo_dispositivo"],
                    entry["dispositivo_nombre"],
                    entry["fecha_evento"].strftime("%d/%m/%Y %H:%M"),
                ]
            )
        bajas_table = Table(bajas_data, hAlign="LEFT")
        bajas_table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor("#b91c1c"),
                    ),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ]
            )
        )
        story.append(bajas_table)

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf


def gather_enterprise_report(
    start: Optional[date],
    end: Optional[date],
    segment: str = SEGMENT_INFORMATICA,
) -> Dict[str, object]:
    if segment not in {SEGMENT_INFORMATICA, SEGMENT_MEDICA}:
        segment = SEGMENT_INFORMATICA

    range_data = default_date_range(start, end)
    start_date = range_data["start"]
    end_date = range_data["end"]
    today = timezone.now().date()

    asset_configs = [
        config for config in ASSET_CONFIGS if config.segment == segment
    ]
    if not asset_configs:
        asset_configs = [
            config
            for config in ASSET_CONFIGS
            if config.segment == SEGMENT_INFORMATICA
        ]

    warranty_summary = {
        "vigentes": 0,
        "por_vencer": 0,
        "vencidas": 0,
        "detalle": [],
    }

    obsolescence_summary = {
        "threshold_years": OBSOLESCENCE_YEARS,
        "total_evaluados": 0,
        "obsoletos": 0,
        "por_tipo": [],
    }

    license_summary = {
        "vigentes": 0,
        "por_vencer": 0,
        "vencidas": 0,
        "detalle": [],
    }

    annual_purchases: List[Dict[str, object]] = []
    monthly_deliveries: List[Dict[str, object]] = []
    annual_deliveries: List[Dict[str, object]] = []

    bajas_summary = {"total_bajas": 0, "detalle": []}

    obsolescence_by_type: Dict[str, Dict[str, float]] = {}

    for config in asset_configs:
        queryset = config.model.objects.all()

        if config.warranty_field:
            field = config.warranty_field
            warranty_filter = {
                f"{field}__isnull": False,
                f"{field}__range": (start_date, end_date),
            }
            for asset in queryset.filter(**warranty_filter).only(
                "nombre", field
            ):
                expiration = getattr(asset, field)
                if not expiration:
                    continue
                estado = "vigente"
                if expiration < today:
                    estado = "vencida"
                    warranty_summary["vencidas"] += 1
                elif expiration <= today + timedelta(
                    days=WARRANTY_SOON_THRESHOLD_DAYS
                ):
                    estado = "por_vencer"
                    warranty_summary["por_vencer"] += 1
                else:
                    warranty_summary["vigentes"] += 1

                warranty_summary["detalle"].append(
                    {
                        "nombre": asset.nombre,
                        "tipo": config.label,
                        "fecha": expiration,
                        "estado": estado,
                        "dias_restantes": (expiration - today).days,
                    }
                )

        if config.license_field:
            field = config.license_field
            license_filter = {
                f"{field}__isnull": False,
                f"{field}__range": (start_date, end_date),
            }
            for asset in queryset.filter(**license_filter).only(
                "nombre", field
            ):
                expiration = getattr(asset, field)
                if not expiration:
                    continue
                estado = "vigente"
                if expiration < today:
                    estado = "vencida"
                    license_summary["vencidas"] += 1
                elif expiration <= today + timedelta(
                    days=WARRANTY_SOON_THRESHOLD_DAYS
                ):
                    estado = "por_vencer"
                    license_summary["por_vencer"] += 1
                else:
                    license_summary["vigentes"] += 1

                license_summary["detalle"].append(
                    {
                        "nombre": asset.nombre,
                        "tipo": config.label,
                        "fecha": expiration,
                        "estado": estado,
                        "dias_restantes": (expiration - today).days,
                    }
                )

        if config.acquisition_field:
            acquisition_field = config.acquisition_field
            acquisition_filter = {
                f"{acquisition_field}__isnull": False,
                f"{acquisition_field}__range": (start_date, end_date),
            }
            acquisition_qs = queryset.filter(**acquisition_filter)
            for asset in acquisition_qs.only("nombre", acquisition_field):
                acquisition_date = getattr(asset, acquisition_field)
                if not acquisition_date:
                    continue
                age_years = (end_date - acquisition_date).days / 365.25
                obsolescence_summary["total_evaluados"] += 1
                entry = obsolescence_by_type.setdefault(
                    config.label,
                    {
                        "total": 0,
                        "obsoletos": 0,
                        "antiguedad_acumulada": 0.0,
                    },
                )
                entry["total"] += 1
                entry["antiguedad_acumulada"] += max(age_years, 0)
                if age_years >= OBSOLESCENCE_YEARS:
                    entry["obsoletos"] += 1
                    obsolescence_summary["obsoletos"] += 1

        if config.value_field and config.acquisition_field:
            value_field = config.value_field
            acquisition_field = config.acquisition_field
            purchase_filter = {
                f"{value_field}__isnull": False,
                f"{value_field}__gt": 0,
                f"{acquisition_field}__range": (start_date, end_date),
            }
            purchases = (
                queryset.filter(**purchase_filter)
                .annotate(year=TruncYear(acquisition_field))
                .values("year")
                .annotate(
                    monto=Sum(value_field),
                    cantidad=Count("id"),
                )
                .order_by("year")
            )
            for purchase in purchases:
                year = (
                    purchase["year"].year if purchase["year"] else "Sin fecha"
                )
                annual_purchases.append(
                    {
                        "tipo": config.label,
                        "year": year,
                        "monto": purchase["monto"] or Decimal(0),
                        "cantidad": purchase["cantidad"],
                    }
                )

    if obsolescence_by_type:
        for label, stats in obsolescence_by_type.items():
            total = stats["total"] or 1
            obsolescence_summary["por_tipo"].append(
                {
                    "tipo": label,
                    "total": stats["total"],
                    "obsoletos": stats["obsoletos"],
                    "porcentaje": round((stats["obsoletos"] / total) * 100, 1),
                    "antiguedad_promedio": round(
                        stats["antiguedad_acumulada"] / total, 1
                    ),
                }
            )

    bitacora_qs = Bitacora.objects.filter(
        fecha_evento__date__range=(start_date, end_date)
    )

    tipos_bitacora = BITACORA_SEGMENT_TYPES.get(segment)
    if tipos_bitacora:
        bitacora_qs = bitacora_qs.filter(
            tipo_dispositivo__in=tipos_bitacora
        )

    entregas_qs = bitacora_qs.filter(tipo_evento="asignacion_personal")
    if entregas_qs.exists():
        monthly = (
            entregas_qs.annotate(month=TruncMonth("fecha_evento"))
            .values("month")
            .annotate(total=Count("id"))
            .order_by("month")
        )
        for entry in monthly:
            monthly_deliveries.append(
                {
                    "periodo": entry["month"].strftime("%Y-%m"),
                    "total": entry["total"],
                }
            )

        annual = (
            entregas_qs.annotate(year=TruncYear("fecha_evento"))
            .values("year")
            .annotate(total=Count("id"))
            .order_by("year")
        )
        for entry in annual:
            annual_deliveries.append(
                {
                    "year": (
                        entry["year"].year if entry["year"] else "Sin fecha"
                    ),
                    "total": entry["total"],
                }
            )

    bajas_qs = bitacora_qs.filter(tipo_evento="baja")
    bajas_summary["total_bajas"] = bajas_qs.count()
    bajas_summary["detalle"] = list(
        bajas_qs.values(
            "tipo_dispositivo",
            "dispositivo_nombre",
            "fecha_evento",
        )
    )

    warranty_summary["detalle"].sort(key=lambda item: item["fecha"])
    license_summary["detalle"].sort(key=lambda item: item["fecha"])
    annual_purchases.sort(key=lambda item: (item["year"], item["tipo"]))

    return {
        "generated_at": timezone.now(),
        "start_date": start_date,
        "end_date": end_date,
        "warranty": warranty_summary,
        "obsolescence": obsolescence_summary,
        "licenses": license_summary,
        "annual_purchases": annual_purchases,
        "monthly_deliveries": monthly_deliveries,
        "annual_deliveries": annual_deliveries,
        "bajas": bajas_summary,
    }
